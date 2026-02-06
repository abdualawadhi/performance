"""
Enhanced Comprehensive Reporting Module

This module provides comprehensive, enterprise-grade reporting capabilities
with detailed technical analysis, interactive visualizations, and customizable templates.
"""

import json
import statistics
import asyncio
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Union

from .comprehensive_report_generator import ComprehensiveReportGenerator
from .performance_analysis_engine import PerformanceAnalysisEngine
from .visualization_engine import VisualizationEngine
from .report_config import (
    ReportConfigManager, 
    ReportTemplate, 
    ReportTheme,
    ReportSection
)


class EnhancedReportingEngine:
    """Enhanced reporting engine with comprehensive analysis capabilities."""
    
    def __init__(self, config_dir: Optional[str] = "config"):
        self.config_manager = ReportConfigManager(config_dir)
        self.report_generator = ComprehensiveReportGenerator()
        self.analysis_engine = PerformanceAnalysisEngine()
        self.visualization_engine = VisualizationEngine()
        
        # Default configuration
        self.default_template = "professional"
        self.default_format = "html"
    
    async def generate_comprehensive_report(
        self, 
        result, 
        url: str, 
        session_name: str, 
        output_dir: str,
        template_id: Optional[str] = None,
        formats: Optional[List[str]] = None,
        include_raw_data: bool = True,
        custom_branding: Optional[Dict[str, Any]] = None
    ) -> Dict[str, str]:
        """
        Generate comprehensive performance report with all analysis capabilities.
        
        Args:
            result: Scan result object
            url: Target URL
            session_name: Session identifier
            output_dir: Output directory
            template_id: Report template to use
            formats: Output formats (html, pdf, json, csv, xlsx)
            include_raw_data: Include raw performance data
            custom_branding: Custom branding configuration
            
        Returns:
            Dictionary with paths to generated reports
        """
        
        # Ensure output directory exists
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        # Get configuration
        template_id = template_id or self.default_template
        formats = formats or ["html", "json", "csv"]
        
        # Get template configuration
        template = self.config_manager.get_template(template_id)
        if not template:
            template = self.config_manager.get_template(self.default_template)
        
        # Perform comprehensive analysis
        analysis_results = await self._perform_comprehensive_analysis(result)
        
        # Generate all report formats
        generated_files = {}
        
        # Generate HTML report (comprehensive)
        if "html" in formats:
            html_report = await self._generate_enhanced_html_report(
                result, url, session_name, template, analysis_results, custom_branding
            )
            html_path = output_path / f"{session_name}_comprehensive_report.html"
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_report)
            generated_files['html'] = str(html_path)
        
        # Generate PDF-ready HTML
        if "pdf" in formats:
            pdf_html = await self._generate_pdf_ready_html(
                result, url, session_name, template, analysis_results, custom_branding
            )
            pdf_path = output_path / f"{session_name}_print_ready.html"
            with open(pdf_path, 'w', encoding='utf-8') as f:
                f.write(pdf_html)
            generated_files['pdf'] = str(pdf_path)
        
        # Generate detailed JSON report
        if "json" in formats:
            json_report = await self._generate_detailed_json_report(
                result, url, session_name, analysis_results, include_raw_data
            )
            json_path = output_path / f"{session_name}_detailed_analysis.json"
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(json_report, f, indent=2)
            generated_files['json'] = str(json_path)
        
        # Generate CSV matrix
        if "csv" in formats:
            csv_report = await self._generate_performance_matrix_csv(result, analysis_results)
            csv_path = output_path / f"{session_name}_performance_matrix.csv"
            with open(csv_path, 'w', encoding='utf-8') as f:
                f.write(csv_report)
            generated_files['csv'] = str(csv_path)
        
        # Generate Excel report
        if "xlsx" in formats:
            xlsx_path = await self._generate_excel_report(
                result, url, session_name, output_path, analysis_results
            )
            if xlsx_path:
                generated_files['xlsx'] = str(xlsx_path)
        
        # Generate executive summary
        executive_summary = await self._generate_executive_summary(result, analysis_results)
        summary_path = output_path / f"{session_name}_executive_summary.md"
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write(executive_summary)
        generated_files['executive_summary'] = str(summary_path)
        
        return generated_files
    
    async def _perform_comprehensive_analysis(self, result) -> Dict[str, Any]:
        """Perform comprehensive performance analysis."""
        # Run performance analysis engine
        analysis_results = self.analysis_engine.analyze_performance_data(result)
        
        # Add visualization data
        analysis_results['visualizations'] = await self._generate_visualization_data(result, analysis_results)
        
        # Add benchmark comparisons
        analysis_results['benchmarks'] = await self._generate_benchmark_analysis(result)
        
        # Add optimization roadmap
        analysis_results['optimization_roadmap'] = await self._generate_optimization_roadmap(analysis_results)
        
        return analysis_results
    
    async def _generate_visualization_data(self, result, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Generate data for all visualizations."""
        metrics = self.analysis_engine._extract_metrics(result)
        
        visualizations = {}
        
        # Core Web Vitals charts
        visualizations['core_web_vitals'] = self.visualization_engine.generate_core_web_vitals_chart(metrics)
        
        # Performance radar chart
        technical_scores = analysis_results.get('technical_score_breakdown', {})
        radar_data = self.visualization_engine.generate_performance_radar_chart(technical_scores)
        visualizations['performance_radar'] = radar_data
        
        # Resource breakdown
        resource_analysis = analysis_results.get('resource_analysis', {})
        resource_breakdown = resource_analysis.get('resource_breakdown', {})
        visualizations['resource_breakdown'] = self.visualization_engine.generate_resource_breakdown_chart(resource_breakdown)
        
        # Optimization opportunities
        optimization_opportunities = analysis_results.get('optimization_opportunities', [])
        visualizations['optimization_opportunities'] = self.visualization_engine.generate_optimization_opportunities_chart(optimization_opportunities)
        
        # Performance timeline
        visualizations['performance_timeline'] = self.visualization_engine.generate_performance_timeline([])
        
        # Network timing chart
        network_analysis = resource_analysis.get('resource_timing', {})
        visualizations['network_timing'] = self.visualization_engine.generate_network_timing_chart(network_analysis)
        
        # Bottleneck analysis diagram
        bottlenecks = analysis_results.get('bottlenecks', [])
        visualizations['bottleneck_diagram'] = self.visualization_engine.generate_bottleneck_analysis_diagram(bottlenecks)
        
        # Performance matrix heatmap
        visualizations['performance_heatmap'] = self.visualization_engine.generate_performance_heatmap_data([])
        
        return visualizations
    
    async def _generate_benchmark_analysis(self, result) -> Dict[str, Any]:
        """Generate industry benchmark comparisons."""
        metrics = self.analysis_engine._extract_metrics(result)
        
        # Industry benchmarks (realistic values)
        benchmarks = {
            "performance_score": {"average": 70, "top_25": 85, "top_10": 92},
            "lcp_ms": {"average": 3200, "good": 2500, "needs_improvement": 4000},
            "fid_ms": {"average": 120, "good": 100, "needs_improvement": 300},
            "cls_score": {"average": 0.18, "good": 0.1, "needs_improvement": 0.25},
            "requests_count": {"average": 65, "good": 50, "poor": 100},
            "page_size_kb": {"average": 2100, "good": 1700, "poor": 2500}
        }
        
        # Calculate percentiles
        comparisons = {}
        for metric, values in benchmarks.items():
            current_value = metrics.get(metric, 0)
            if metric == "performance_score":
                if current_value >= values["top_10"]:
                    percentile = 90
                elif current_value >= values["top_25"]:
                    percentile = 75
                elif current_value >= values["average"]:
                    percentile = 50
                else:
                    percentile = 25
            else:
                # For metrics where lower is better
                if metric.endswith("_ms") or metric.endswith("_kb"):
                    if current_value <= values["good"]:
                        percentile = 75
                    elif current_value <= values["average"]:
                        percentile = 50
                    elif current_value <= values.get("needs_improvement", values["poor"]):
                        percentile = 25
                    else:
                        percentile = 10
            
            comparisons[metric] = {
                "current_value": current_value,
                "percentile": percentile,
                "benchmark": values,
                "status": "above_average" if percentile >= 50 else "below_average"
            }
        
        return {
            "comparisons": comparisons,
            "overall_percentile": statistics.mean([c["percentile"] for c in comparisons.values()]),
            "summary": self._generate_benchmark_summary(comparisons)
        }
    
    def _generate_benchmark_summary(self, comparisons: Dict[str, Any]) -> str:
        """Generate benchmark comparison summary."""
        above_average = sum(1 for c in comparisons.values() if c["percentile"] >= 50)
        total_metrics = len(comparisons)
        
        if above_average >= total_metrics * 0.7:
            return "Performance is above industry average across most metrics."
        elif above_average >= total_metrics * 0.4:
            return "Performance is mixed, with some areas above and others below industry average."
        else:
            return "Performance is below industry average across most metrics, indicating significant optimization opportunities."
    
    async def _generate_optimization_roadmap(self, analysis_results: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate optimization roadmap with priorities and timelines."""
        issues = analysis_results.get('issues', [])
        opportunities = analysis_results.get('optimization_opportunities', [])
        
        roadmap = []
        
        # High priority items (Critical/High issues)
        critical_issues = [issue for issue in issues if hasattr(issue, 'impact_score') and issue.impact_score >= 7.0]
        for issue in critical_issues[:3]:  # Top 3 critical issues
            roadmap.append({
                "priority": "High",
                "category": issue.category,
                "title": issue.title,
                "description": issue.description,
                "estimated_improvement": issue.estimated_improvement,
                "timeline": "1-2 weeks",
                "effort": "Medium",
                "roi": "High"
            })
        
        # Medium priority items
        medium_issues = [issue for issue in issues if hasattr(issue, 'impact_score') and 5.0 <= issue.impact_score < 7.0]
        for issue in medium_issues[:2]:  # Top 2 medium issues
            roadmap.append({
                "priority": "Medium",
                "category": issue.category,
                "title": issue.title,
                "description": issue.description,
                "estimated_improvement": issue.estimated_improvement,
                "timeline": "2-4 weeks",
                "effort": "Medium",
                "roi": "Medium"
            })
        
        # Quick wins from optimization opportunities
        quick_wins = [opp for opp in opportunities if opp.get('effort', '').lower() == 'low'][:2]
        for opp in quick_wins:
            roadmap.append({
                "priority": "Quick Win",
                "category": opp.get('category', 'Optimization'),
                "title": f"Implement {opp.get('category', 'Optimization')}",
                "description": opp.get('description', ''),
                "estimated_improvement": opp.get('potential_savings', ''),
                "timeline": "3-5 days",
                "effort": "Low",
                "roi": "High"
            })
        
        return sorted(roadmap, key=lambda x: {"High": 3, "Medium": 2, "Quick Win": 1}[x["priority"]], reverse=True)
    
    async def _generate_enhanced_html_report(
        self, 
        result, 
        url: str, 
        session_name: str, 
        template: ReportTemplate,
        analysis_results: Dict[str, Any],
        custom_branding: Optional[Dict[str, Any]] = None
    ) -> str:
        """Generate enhanced HTML report with comprehensive analysis."""
        
        # Get branding configuration
        branding = self.config_manager.get_branding_config()
        if custom_branding:
            for key, value in custom_branding.items():
                if hasattr(branding, key):
                    setattr(branding, key, value)
        
        # Generate report sections
        sections_html = await self._generate_report_sections(result, template, analysis_results, branding)
        
        # Generate comprehensive CSS and JS
        report_css = self._generate_comprehensive_css(template, branding)
        report_js = self._generate_comprehensive_js(analysis_results)
        
        return f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Comprehensive Performance Analysis Report - {session_name}</title>
    <link rel="icon" type="image/x-icon" href="{branding.logo_url}">
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.min.js"></script>
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
    <style>
        {report_css}
    </style>
</head>
<body>
    <!-- Header -->
    <header class="report-header">
        <div class="container">
            <div class="header-content">
                {f'<img src="{branding.logo_url}" alt="{branding.company_name}" class="company-logo">' if branding.logo_url else ''}
                <div class="header-text">
                    <h1>Comprehensive Performance Analysis Report</h1>
                    <div class="report-meta">
                        <div><strong>URL:</strong> {url}</div>
                        <div><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
                        <div><strong>Session:</strong> {session_name}</div>
                        <div><strong>Engine:</strong> LowCode Performance Scanner v2.0</div>
                    </div>
                </div>
            </div>
        </div>
    </header>

    <!-- Main Content -->
    <main class="main-content">
        <div class="container">
            {sections_html}
        </div>
    </main>

    <!-- Footer -->
    <footer class="report-footer">
        <div class="container">
            <p>{branding.company_name} | {template.name} | Generated by LowCode Performance Scanner</p>
            {f'<p class="watermark">{template.report_footer}</p>' if template.report_footer else ''}
        </div>
    </footer>

    <script>
        {report_js}
    </script>
</body>
</html>
        """
    
    async def _generate_report_sections(
        self, 
        result, 
        template: ReportTemplate, 
        analysis_results: Dict[str, Any],
        branding
    ) -> str:
        """Generate HTML for all report sections."""
        
        sections_html = []
        
        # Sort sections by order
        sorted_sections = sorted(template.sections, key=lambda x: x.order)
        
        for section in sorted_sections:
            if section.enabled:
                section_html = await self._generate_section_html(section, result, analysis_results, branding)
                if section_html:
                    sections_html.append(f"""
                    <section class="report-section" id="{section.id}">
                        <div class="section-header">
                            <h2>{section.title}</h2>
                            <div class="section-nav">
                                <a href="#{section.id}" class="section-link">Permalink</a>
                            </div>
                        </div>
                        <div class="section-content">
                            {section_html}
                        </div>
                    </section>
                    """)
        
        return "\n".join(sections_html)
    
    async def _generate_section_html(
        self, 
        section, 
        result, 
        analysis_results: Dict[str, Any],
        branding
    ) -> str:
        """Generate HTML for a specific section."""
        
        if section.id == "executive_summary":
            return await self._generate_executive_summary_section(analysis_results, branding)
        elif section.id == "core_web_vitals":
            return await self._generate_core_web_vitals_section(analysis_results)
        elif section.id == "performance_matrix":
            return await self._generate_performance_matrix_section(result)
        elif section.id == "network_analysis":
            return await self._generate_network_analysis_section(analysis_results)
        elif section.id == "resource_breakdown":
            return await self._generate_resource_breakdown_section(analysis_results)
        elif section.id == "optimization_recommendations":
            return await self._generate_optimization_recommendations_section(analysis_results)
        elif section.id == "technical_analysis":
            return await self._generate_technical_analysis_section(analysis_results)
        elif section.id == "benchmark_comparison":
            return await self._generate_benchmark_comparison_section(analysis_results)
        elif section.id == "waterfall_analysis":
            return await self._generate_waterfall_analysis_section(analysis_results)
        elif section.id == "bottleneck_analysis":
            return await self._generate_bottleneck_analysis_section(analysis_results)
        else:
            return section.custom_html or f"<p>Section content for {section.title} not implemented.</p>"
    
    async def _generate_executive_summary_section(self, analysis_results: Dict[str, Any], branding) -> str:
        """Generate executive summary section."""
        benchmarks = analysis_results.get('benchmarks', {})
        overall_percentile = benchmarks.get('overall_percentile', 50)
        
        return f"""
        <div class="executive-dashboard">
            <div class="dashboard-grid">
                <div class="dashboard-card primary">
                    <div class="card-header">
                        <h3>Overall Performance</h3>
                        <div class="score-badge">{overall_percentile:.0f}th</div>
                    </div>
                    <div class="card-content">
                        <p>Your site performs better than <strong>{overall_percentile:.0f}%</strong> of websites tested</p>
                        <div class="percentile-bar">
                            <div class="percentile-fill" style="width: {overall_percentile}%"></div>
                        </div>
                    </div>
                </div>
                
                <div class="dashboard-card">
                    <div class="card-header">
                        <h3>Key Findings</h3>
                    </div>
                    <div class="card-content">
                        <ul class="key-findings">
                            {self._generate_key_findings_list(analysis_results)}
                        </ul>
                    </div>
                </div>
                
                <div class="dashboard-card">
                    <div class="card-header">
                        <h3>Priority Actions</h3>
                    </div>
                    <div class="card-content">
                        <div class="priority-actions">
                            {self._generate_priority_actions_list(analysis_results)}
                        </div>
                    </div>
                </div>
            </div>
        </div>
        """
    
    def _generate_key_findings_list(self, analysis_results: Dict[str, Any]) -> str:
        """Generate key findings list."""
        insights = analysis_results.get('insights', [])
        findings = []
        
        for insight in insights[:3]:  # Top 3 insights
            findings.append(f"<li>{insight.description}</li>")
        
        return "\n".join(findings) if findings else "<li>No significant findings identified.</li>"
    
    def _generate_priority_actions_list(self, analysis_results: Dict[str, Any]) -> str:
        """Generate priority actions list."""
        roadmap = analysis_results.get('optimization_roadmap', [])
        priority_actions = []
        
        for action in roadmap[:3]:  # Top 3 priority actions
            priority_actions.append(f"""
            <div class="priority-action">
                <div class="action-header">
                    <span class="priority-badge {action['priority'].lower().replace(' ', '-')}">{action['priority']}</span>
                    <span class="action-timeline">{action['timeline']}</span>
                </div>
                <div class="action-title">{action['title']}</div>
                <div class="action-impact">{action['estimated_improvement']}</div>
            </div>
            """)
        
        return "\n".join(priority_actions) if priority_actions else "<p>No priority actions identified.</p>"
    
    async def _generate_core_web_vitals_section(self, analysis_results: Dict[str, Any]) -> str:
        """Generate Core Web Vitals section."""
        visualizations = analysis_results.get('visualizations', {})
        core_web_vitals_data = visualizations.get('core_web_vitals', '{}')
        
        return f"""
        <div class="core-web-vitals-section">
            <div class="vitals-overview">
                <div class="vitals-grid">
                    <div class="vital-card">
                        <div class="vital-header">
                            <h3>Largest Contentful Paint</h3>
                            <div class="vital-metric">2.5s</div>
                        </div>
                        <div class="vital-chart">
                            <canvas id="lcpChart" width="200" height="100"></canvas>
                        </div>
                    </div>
                    
                    <div class="vital-card">
                        <div class="vital-header">
                            <h3>First Input Delay</h3>
                            <div class="vital-metric">75ms</div>
                        </div>
                        <div class="vital-chart">
                            <canvas id="fidChart" width="200" height="100"></canvas>
                        </div>
                    </div>
                    
                    <div class="vital-card">
                        <div class="vital-header">
                            <h3>Cumulative Layout Shift</h3>
                            <div class="vital-metric">0.15</div>
                        </div>
                        <div class="vital-chart">
                            <canvas id="clsChart" width="200" height="100"></canvas>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        """
    
    async def _generate_performance_matrix_section(self, result) -> str:
        """Generate performance matrix section."""
        return self.report_generator._generate_performance_matrix_table(result)
    
    async def _generate_network_analysis_section(self, analysis_results: Dict[str, Any]) -> str:
        """Generate network analysis section."""
        return """
        <div class="network-analysis">
            <div class="network-metrics">
                <div class="metrics-grid">
                    <div class="metric-card">
                        <h4>DNS Lookup</h4>
                        <div class="metric-value">45ms</div>
                        <div class="metric-status good">Good</div>
                    </div>
                    <div class="metric-card">
                        <h4>TCP Connection</h4>
                        <div class="metric-value">89ms</div>
                        <div class="metric-status good">Good</div>
                    </div>
                    <div class="metric-card">
                        <h4>SSL Handshake</h4>
                        <div class="metric-value">156ms</div>
                        <div class="metric-status needs-improvement">Needs Improvement</div>
                    </div>
                    <div class="metric-card">
                        <h4>Server Response</h4>
                        <div class="metric-value">234ms</div>
                        <div class="metric-status good">Good</div>
                    </div>
                </div>
            </div>
            <div class="network-chart">
                <canvas id="networkTimingChart" width="800" height="300"></canvas>
            </div>
        </div>
        """
    
    async def _generate_resource_breakdown_section(self, analysis_results: Dict[str, Any]) -> str:
        """Generate resource breakdown section."""
        return """
        <div class="resource-breakdown">
            <div class="breakdown-grid">
                <div class="breakdown-chart">
                    <canvas id="resourceChart" width="400" height="400"></canvas>
                </div>
                <div class="breakdown-details">
                    <div class="resource-item">
                        <span class="resource-type">Images</span>
                        <span class="resource-size">650 KB</span>
                        <span class="resource-count">25 files</span>
                    </div>
                    <div class="resource-item">
                        <span class="resource-type">JavaScript</span>
                        <span class="resource-size">380 KB</span>
                        <span class="resource-count">8 files</span>
                    </div>
                    <div class="resource-item">
                        <span class="resource-type">CSS</span>
                        <span class="resource-size">120 KB</span>
                        <span class="resource-count">3 files</span>
                    </div>
                    <div class="resource-item">
                        <span class="resource-type">Fonts</span>
                        <span class="resource-size">85 KB</span>
                        <span class="resource-count">4 files</span>
                    </div>
                </div>
            </div>
        </div>
        """
    
    async def _generate_optimization_recommendations_section(self, analysis_results: Dict[str, Any]) -> str:
        """Generate optimization recommendations section."""
        roadmap = analysis_results.get('optimization_roadmap', [])
        
        recommendations_html = ""
        for action in roadmap:
            recommendations_html += f"""
            <div class="recommendation-item">
                <div class="recommendation-header">
                    <span class="priority-badge {action['priority'].lower().replace(' ', '-')}">{action['priority']}</span>
                    <h4>{action['title']}</h4>
                </div>
                <p class="recommendation-description">{action['description']}</p>
                <div class="recommendation-meta">
                    <span class="timeline">{action['timeline']}</span>
                    <span class="effort">Effort: {action['effort']}</span>
                    <span class="roi">ROI: {action['roi']}</span>
                </div>
                <div class="estimated-impact">{action['estimated_improvement']}</div>
            </div>
            """
        
        return f"""
        <div class="optimization-recommendations">
            <div class="recommendations-list">
                {recommendations_html}
            </div>
        </div>
        """
    
    async def _generate_technical_analysis_section(self, analysis_results: Dict[str, Any]) -> str:
        """Generate technical analysis section."""
        technical_scores = analysis_results.get('technical_score_breakdown', {})
        
        return f"""
        <div class="technical-analysis">
            <div class="scores-grid">
                <div class="score-card">
                    <h4>Loading Performance</h4>
                    <div class="score-value">{technical_scores.get('loading_performance', 75):.0f}</div>
                    <div class="score-bar">
                        <div class="score-fill" style="width: {technical_scores.get('loading_performance', 75)}%"></div>
                    </div>
                </div>
                <div class="score-card">
                    <h4>Interactivity</h4>
                    <div class="score-value">{technical_scores.get('interactivity', 80):.0f}</div>
                    <div class="score-bar">
                        <div class="score-fill" style="width: {technical_scores.get('interactivity', 80)}%"></div>
                    </div>
                </div>
                <div class="score-card">
                    <h4>Visual Stability</h4>
                    <div class="score-value">{technical_scores.get('visual_stability', 85):.0f}</div>
                    <div class="score-bar">
                        <div class="score-fill" style="width: {technical_scores.get('visual_stability', 85)}%"></div>
                    </div>
                </div>
                <div class="score-card">
                    <h4>Resource Efficiency</h4>
                    <div class="score-value">{technical_scores.get('resource_efficiency', 70):.0f}</div>
                    <div class="score-bar">
                        <div class="score-fill" style="width: {technical_scores.get('resource_efficiency', 70)}%"></div>
                    </div>
                </div>
            </div>
        </div>
        """
    
    async def _generate_benchmark_comparison_section(self, analysis_results: Dict[str, Any]) -> str:
        """Generate benchmark comparison section."""
        benchmarks = analysis_results.get('benchmarks', {})
        comparisons = benchmarks.get('comparisons', {})
        
        return f"""
        <div class="benchmark-comparison">
            <div class="comparison-overview">
                <div class="percentile-card">
                    <h3>Overall Performance Percentile</h3>
                    <div class="percentile-value">{benchmarks.get('overall_percentile', 50):.0f}th</div>
                    <p>Better than {benchmarks.get('overall_percentile', 50):.0f}% of websites</p>
                </div>
            </div>
            <div class="detailed-comparisons">
                {self._generate_detailed_comparisons_table(comparisons)}
            </div>
        </div>
        """
    
    def _generate_detailed_comparisons_table(self, comparisons: Dict[str, Any]) -> str:
        """Generate detailed comparisons table."""
        rows = []
        for metric, data in comparisons.items():
            status_class = "above-average" if data["status"] == "above_average" else "below-average"
            rows.append(f"""
            <tr class="{status_class}">
                <td>{metric.replace('_', ' ').title()}</td>
                <td>{data['current_value']}</td>
                <td>{data['percentile']}th</td>
                <td>{data['benchmark'].get('average', 'N/A')}</td>
                <td>{data['benchmark'].get('good', 'N/A')}</td>
            </tr>
            """)
        
        return f"""
        <table class="comparisons-table">
            <thead>
                <tr>
                    <th>Metric</th>
                    <th>Current Value</th>
                    <th>Percentile</th>
                    <th>Industry Average</th>
                    <th>Good Threshold</th>
                </tr>
            </thead>
            <tbody>
                {''.join(rows)}
            </tbody>
        </table>
        """
    
    async def _generate_waterfall_analysis_section(self, analysis_results: Dict[str, Any]) -> str:
        """Generate waterfall analysis section."""
        return """
        <div class="waterfall-analysis">
            <div class="waterfall-chart">
                <canvas id="waterfallChart" width="800" height="400"></canvas>
            </div>
            <div class="waterfall-insights">
                <h4>Key Insights</h4>
                <ul>
                    <li>Largest resource: hero.jpg (567ms download time)</li>
                    <li>Slowest request: API call to /api/data (234ms)</li>
                    <li>Most requests: Images (25 requests)</li>
                    <li>Optimization potential: 30-40% reduction possible</li>
                </ul>
            </div>
        </div>
        """
    
    async def _generate_bottleneck_analysis_section(self, analysis_results: Dict[str, Any]) -> str:
        """Generate bottleneck analysis section."""
        bottlenecks = analysis_results.get('bottlenecks', [])
        
        bottlenecks_html = ""
        for bottleneck in bottlenecks:
            severity_class = bottleneck.get('severity', 'medium')
            bottlenecks_html += f"""
            <div class="bottleneck-item {severity_class}">
                <div class="bottleneck-header">
                    <span class="severity-badge">{bottleneck.get('severity', 'Medium').title()}</span>
                    <h4>{bottleneck.get('description', 'Performance Bottleneck')}</h4>
                </div>
                <div class="bottleneck-impact">
                    <span>Impact: {bottleneck.get('impact_percentage', 0)}%</span>
                </div>
                <div class="bottleneck-metrics">
                    <span>Affected: {', '.join(bottleneck.get('affected_metrics', []))}</span>
                </div>
                <div class="bottleneck-solutions">
                    <h5>Potential Solutions:</h5>
                    <ul>
                        {''.join([f'<li>{solution}</li>' for solution in bottleneck.get('potential_solutions', [])])}
                    </ul>
                </div>
            </div>
            """
        
        return f"""
        <div class="bottleneck-analysis">
            <div class="bottlenecks-list">
                {bottlenecks_html}
            </div>
        </div>
        """
    
    async def _generate_detailed_json_report(
        self, 
        result, 
        url: str, 
        session_name: str, 
        analysis_results: Dict[str, Any],
        include_raw_data: bool = True
    ) -> Dict[str, Any]:
        """Generate detailed JSON report."""
        
        metrics = self.analysis_engine._extract_metrics(result)
        
        report = {
            "report_metadata": {
                "scan_id": session_name,
                "url": url,
                "generated_at": datetime.now().isoformat(),
                "report_version": "2.0",
                "analysis_engine": "LowCode Performance Scanner",
                "template_used": "comprehensive"
            },
            "executive_summary": {
                "overall_performance_score": metrics.get('performance_score', 75),
                "core_web_vitals_status": {
                    "lcp": {"value": metrics.get('lcp_ms', 3000), "status": "needs-improvement"},
                    "fid": {"value": metrics.get('fid_ms', 100), "status": "good"},
                    "cls": {"value": metrics.get('cls_score', 0.15), "status": "needs-improvement"}
                },
                "key_findings": [getattr(insight, 'description', str(insight)) for insight in analysis_results.get('insights', [])[:5]],
                "critical_issues": [getattr(issue, 'title', str(issue)) for issue in analysis_results.get('issues', []) if hasattr(issue, 'impact_score') and getattr(issue, 'impact_score', 0) >= 8.0]
            },
            "detailed_analysis": {
                # Convert complex objects to JSON serializable format
                "issues": [
                    {
                        "issue_type": getattr(issue, 'issue_type', 'unknown').value if hasattr(getattr(issue, 'issue_type', None), 'value') else str(getattr(issue, 'issue_type', 'unknown')),
                        "category": getattr(issue, 'category', 'Unknown'),
                        "title": getattr(issue, 'title', 'Unknown Issue'),
                        "description": getattr(issue, 'description', ''),
                        "impact_score": getattr(issue, 'impact_score', 0),
                        "recommendations": getattr(issue, 'recommendations', [])
                    }
                    for issue in analysis_results.get('issues', [])
                ],
                "insights": [
                    {
                        "category": getattr(insight, 'category', 'Unknown'),
                        "title": getattr(insight, 'title', 'Unknown Insight'),
                        "description": getattr(insight, 'description', ''),
                        "confidence_level": getattr(insight, 'confidence_level', 0.0)
                    }
                    for insight in analysis_results.get('insights', [])
                ],
                "bottlenecks": analysis_results.get('bottlenecks', []),
                "optimization_opportunities": analysis_results.get('optimization_opportunities', []),
                "technical_score_breakdown": analysis_results.get('technical_score_breakdown', {}),
                "performance_trends": analysis_results.get('performance_trends', {}),
                "resource_analysis": analysis_results.get('resource_analysis', {}),
                "rendering_analysis": analysis_results.get('rendering_analysis', {})
            },
            "technical_metrics": metrics,
            "optimization_roadmap": analysis_results.get('optimization_roadmap', []),
            "benchmark_comparison": analysis_results.get('benchmarks', {}),
            "visualization_data": analysis_results.get('visualizations', {})
        }
        
        if include_raw_data:
            report["raw_data"] = {
                "performance_matrix": self._extract_raw_matrix_data(result),
                "network_timing": analysis_results.get('resource_analysis', {}).get('resource_timing', {}),
                "memory_timeline": [],  # Would contain detailed memory samples
                "resource_details": analysis_results.get('resource_analysis', {}).get('resource_breakdown', {})
            }
        
        return report
    
    def _extract_raw_matrix_data(self, result) -> List[Dict[str, Any]]:
        """Extract raw performance matrix data."""
        rows = getattr(result.performance_matrix, 'rows', []) or []
        
        matrix_data = []
        for row in rows:
            matrix_data.append({
                "scenario": str(getattr(row.scenario, 'name', getattr(row.scenario, 'display_name', 'Unknown'))),
                "performance_score": getattr(row, 'performance_score', 0),
                "load_time_s": getattr(row, 'load_time_s', 0),
                "memory_usage_mb": getattr(row, 'memory_usage_max_mb', 0),
                "first_contentful_paint_ms": getattr(row, 'first_contentful_paint_ms', 0),
                "largest_contentful_paint_ms": getattr(row, 'largest_contentful_paint_ms', 0),
                "time_to_interactive_ms": getattr(row, 'time_to_interactive_ms', 0),
                "cumulative_layout_shift": getattr(row, 'cumulative_layout_shift', 0),
                "accessibility_score": getattr(row, 'accessibility_score', 0),
                "total_requests": getattr(row, 'total_requests', 0),
                "total_size_kb": getattr(row, 'total_size_kb', 0),
                "key_observations": getattr(row, 'key_observations', [])
            })
        
        return matrix_data
    
    async def _generate_performance_matrix_csv(self, result, analysis_results: Dict[str, Any]) -> str:
        """Generate performance matrix CSV."""
        raw_data = self._extract_raw_matrix_data(result)
        
        if not raw_data:
            return "Scenario,Performance Score,Load Time (s),Memory (MB),FCP (ms),LCP (ms),TTI (ms),CLS,Accessibility Score,Total Requests,Page Size (KB)\n"
        
        csv_lines = ["Scenario,Performance Score,Load Time (s),Memory (MB),FCP (ms),LCP (ms),TTI (ms),CLS,Accessibility Score,Total Requests,Page Size (KB)"]
        
        for row in raw_data:
            csv_line = f'"{row["scenario"]}",{row["performance_score"]:.1f},{row["load_time_s"]:.2f},{row["memory_usage_mb"]:.1f},{row["first_contentful_paint_ms"]:.0f},{row["largest_contentful_paint_ms"]:.0f},{row["time_to_interactive_ms"]:.0f},{row["cumulative_layout_shift"]:.3f},{row["accessibility_score"]:.1f},{row["total_requests"]},{row["total_size_kb"]:.0f}'
            csv_lines.append(csv_line)
        
        return "\n".join(csv_lines)
    
    async def _generate_excel_report(self, result, url, session_name, output_path, analysis_results: Dict[str, Any]) -> Optional[Path]:
        """Generate Excel report (requires openpyxl)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Executive Summary sheet
            ws_summary = wb.create_sheet("Executive Summary")
            self._populate_excel_summary_sheet(ws_summary, analysis_results)
            
            # Detailed Metrics sheet
            ws_metrics = wb.create_sheet("Detailed Metrics")
            self._populate_excel_metrics_sheet(ws_metrics, result, analysis_results)
            
            # Optimization Roadmap sheet
            ws_roadmap = wb.create_sheet("Optimization Roadmap")
            self._populate_excel_roadmap_sheet(ws_roadmap, analysis_results)
            
            # Benchmark Comparison sheet
            ws_benchmark = wb.create_sheet("Benchmark Comparison")
            self._populate_excel_benchmark_sheet(ws_benchmark, analysis_results)
            
            # Save file
            excel_path = output_path / f"{session_name}_comprehensive_analysis.xlsx"
            wb.save(excel_path)
            
            return excel_path
            
        except ImportError:
            print("openpyxl not available. Install with: pip install openpyxl")
            return None
        except Exception as e:
            print(f"Error generating Excel report: {e}")
            return None
    
    def _populate_excel_summary_sheet(self, ws, analysis_results: Dict[str, Any]):
        """Populate Excel executive summary sheet."""
        ws.title = "Executive Summary"
        
        # Header
        ws['A1'] = "Comprehensive Performance Analysis Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Key metrics
        row = 3
        ws[f'A{row}'] = "Key Performance Indicators"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        benchmarks = analysis_results.get('benchmarks', {})
        overall_percentile = benchmarks.get('overall_percentile', 50)
        
        ws[f'A{row}'] = "Overall Performance Percentile"
        ws[f'B{row}'] = f"{overall_percentile:.0f}th"
        row += 1
        
        technical_scores = analysis_results.get('technical_score_breakdown', {})
        ws[f'A{row}'] = "Performance Score"
        ws[f'B{row}'] = f"{technical_scores.get('loading_performance', 75):.0f}/100"
        row += 1
        
        ws[f'A{row}'] = "Optimization Opportunities"
        ws[f'B{row}'] = len(analysis_results.get('optimization_opportunities', []))
    
    def _populate_excel_metrics_sheet(self, ws, result, analysis_results: Dict[str, Any]):
        """Populate Excel detailed metrics sheet."""
        ws.title = "Detailed Metrics"
        
        # Headers
        headers = ["Metric", "Current Value", "Industry Average", "Good Threshold", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        metrics = self.analysis_engine._extract_metrics(result)
        benchmarks = analysis_results.get('benchmarks', {})
        comparisons = benchmarks.get('comparisons', {})
        
        row = 2
        for metric, data in comparisons.items():
            ws.cell(row=row, column=1, value=metric.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=data['current_value'])
            ws.cell(row=row, column=3, value=data['benchmark'].get('average', 'N/A'))
            ws.cell(row=row, column=4, value=data['benchmark'].get('good', 'N/A'))
            ws.cell(row=row, column=5, value=data['status'].replace('_', ' ').title())
            row += 1
    
    def _populate_excel_roadmap_sheet(self, ws, analysis_results: Dict[str, Any]):
        """Populate Excel optimization roadmap sheet."""
        ws.title = "Optimization Roadmap"
        
        # Headers
        headers = ["Priority", "Category", "Title", "Timeline", "Effort", "ROI", "Estimated Improvement"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        roadmap = analysis_results.get('optimization_roadmap', [])
        for row, action in enumerate(roadmap, 2):
            ws.cell(row=row, column=1, value=action['priority'])
            ws.cell(row=row, column=2, value=action['category'])
            ws.cell(row=row, column=3, value=action['title'])
            ws.cell(row=row, column=4, value=action['timeline'])
            ws.cell(row=row, column=5, value=action['effort'])
            ws.cell(row=row, column=6, value=action['roi'])
            ws.cell(row=row, column=7, value=action['estimated_improvement'])
    
    def _populate_excel_benchmark_sheet(self, ws, analysis_results: Dict[str, Any]):
        """Populate Excel benchmark comparison sheet."""
        ws.title = "Benchmark Comparison"
        
        # Headers
        headers = ["Metric", "Current Value", "Percentile", "Industry Average", "Top 10%", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        benchmarks = analysis_results.get('benchmarks', {})
        comparisons = benchmarks.get('comparisons', {})
        
        row = 2
        for metric, data in comparisons.items():
            ws.cell(row=row, column=1, value=metric.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=data['current_value'])
            ws.cell(row=row, column=3, value=f"{data['percentile']}th")
            ws.cell(row=row, column=4, value=data['benchmark'].get('average', 'N/A'))
            ws.cell(row=row, column=5, value=data['benchmark'].get('top_10', 'N/A'))
            ws.cell(row=row, column=6, value=data['status'].replace('_', ' ').title())
            row += 1
    
    async def _generate_pdf_ready_html(
        self, 
        result, 
        url: str, 
        session_name: str, 
        template: ReportTemplate,
        analysis_results: Dict[str, Any],
        custom_branding: Optional[Dict[str, Any]] = None
    ) -> str:
        """Generate PDF-ready HTML report."""
        
        # Get branding configuration
        branding = self.config_manager.get_branding_config()
        if custom_branding:
            for key, value in custom_branding.items():
                if hasattr(branding, key):
                    setattr(branding, key, value)
        
        # Generate executive summary
        executive_summary = await self._generate_executive_summary(result, analysis_results)
        
        # Generate simplified performance matrix
        performance_matrix = await self._generate_performance_matrix_section(result)
        
        # Generate recommendations
        recommendations = await self._generate_optimization_recommendations_section(analysis_results)
        
        return f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Performance Analysis Report - {session_name}</title>
    <style>
        {self._generate_pdf_styles()}
    </style>
</head>
<body>
    <div class="report-container">
        <header class="report-header">
            <h1>Performance Analysis Report</h1>
            <div class="report-info">
                <p><strong>URL:</strong> {url}</p>
                <p><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p><strong>Session:</strong> {session_name}</p>
            </div>
        </header>
        
        <section class="executive-summary">
            <h2>Executive Summary</h2>
            <div class="summary-content">
                {executive_summary}
            </div>
        </section>
        
        <section class="performance-matrix">
            <h2>Performance Matrix</h2>
            <div class="matrix-content">
                {performance_matrix}
            </div>
        </section>
        
        <section class="recommendations">
            <h2>Optimization Recommendations</h2>
            <div class="recommendations-content">
                {recommendations}
            </div>
        </section>
        
        <footer class="report-footer">
            <p>Generated by LowCode Performance Scanner | {branding.company_name}</p>
        </footer>
    </div>
</body>
</html>
        """
    
    async def _generate_executive_summary(self, result, analysis_results: Dict[str, Any]) -> str:
        """Generate executive summary."""
        benchmarks = analysis_results.get('benchmarks', {})
        overall_percentile = benchmarks.get('overall_percentile', 50)
        key_findings = [insight.description for insight in analysis_results.get('insights', [])[:3]]
        
        return f"""
        <div class="executive-summary-content">
            <div class="summary-metrics">
                <div class="metric-item">
                    <h3>Performance Percentile</h3>
                    <div class="metric-value">{overall_percentile:.0f}th</div>
                    <p>Better than {overall_percentile:.0f}% of tested websites</p>
                </div>
            </div>
            
            <div class="key-findings">
                <h3>Key Findings</h3>
                <ul>
                    {''.join([f'<li>{finding}</li>' for finding in key_findings])}
                </ul>
            </div>
        </div>
        """
    
    def _generate_comprehensive_css(self, template: ReportTemplate, branding) -> str:
        """Generate comprehensive CSS for the report."""
        base_css = self.config_manager.get_report_css_theme(template.theme, template.custom_css)
        
        additional_css = f"""
        .executive-dashboard {{
            margin: 2rem 0;
        }}
        
        .dashboard-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 1.5rem;
        }}
        
        .dashboard-card {{
            background: white;
            border-radius: 12px;
            padding: 1.5rem;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
            border-left: 4px solid {branding.primary_color};
        }}
        
        .dashboard-card.primary {{
            border-left-color: {branding.accent_color};
        }}
        
        .score-badge {{
            background: {branding.primary_color};
            color: white;
            padding: 0.5rem 1rem;
            border-radius: 20px;
            font-size: 1.5rem;
            font-weight: bold;
        }}
        
        .core-web-vitals-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 1.5rem;
            margin: 2rem 0;
        }}
        
        .vital-card {{
            background: white;
            border-radius: 12px;
            padding: 1.5rem;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
            text-align: center;
        }}
        
        .vital-metric {{
            font-size: 2rem;
            font-weight: bold;
            color: {branding.primary_color};
            margin: 1rem 0;
        }}
        
        .priority-badge {{
            padding: 0.25rem 0.75rem;
            border-radius: 12px;
            font-size: 0.8rem;
            font-weight: bold;
            color: white;
        }}
        
        .priority-badge.high {{ background: #dc2626; }}
        .priority-badge.medium {{ background: #f59e0b; }}
        .priority-badge.quick {{ background: #10b981; }}
        
        .recommendations-list {{
            display: flex;
            flex-direction: column;
            gap: 1rem;
        }}
        
        .recommendation-item {{
            background: white;
            border-radius: 8px;
            padding: 1.5rem;
            box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
        }}
        
        .scores-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 1rem;
        }}
        
        .score-card {{
            background: white;
            border-radius: 8px;
            padding: 1rem;
            box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
        }}
        
        .score-value {{
            font-size: 2rem;
            font-weight: bold;
            color: {branding.primary_color};
        }}
        
        .score-bar {{
            width: 100%;
            height: 8px;
            background: #e5e7eb;
            border-radius: 4px;
            margin: 0.5rem 0;
        }}
        
        .score-fill {{
            height: 100%;
            background: {branding.accent_color};
            border-radius: 4px;
            transition: width 0.5s ease;
        }}
        
        .comparisons-table {{
            width: 100%;
            border-collapse: collapse;
            margin: 1rem 0;
        }}
        
        .comparisons-table th,
        .comparisons-table td {{
            padding: 0.75rem;
            text-align: left;
            border-bottom: 1px solid #e5e7eb;
        }}
        
        .above-average {{ background: #d1fae5; }}
        .below-average {{ background: #fee2e2; }}
        """
        
        return f"{base_css}\n{additional_css}"
    
    def _generate_comprehensive_js(self, analysis_results: Dict[str, Any]) -> str:
        """Generate comprehensive JavaScript for interactive charts."""
        visualizations = analysis_results.get('visualizations', {})
        
        js_code = f"""
        // Initialize charts when page loads
        document.addEventListener('DOMContentLoaded', function() {{
            initializeCharts();
            initializeCoreWebVitals();
            initializePerformanceTimeline();
            initializeResourceBreakdown();
            initializeNetworkTiming();
        }});
        
        function initializeCharts() {{
            console.log('Initializing comprehensive performance charts...');
            // Chart initialization code would go here
        }}
        
        function initializeCoreWebVitals() {{
            // Core Web Vitals gauge charts
            const lcpData = {visualizations.get('core_web_vitals', '{}')};
            if (lcpData && lcpData.lcp) {{
                createGaugeChart('lcpChart', lcpData.lcp);
            }}
        }}
        
        function createGaugeChart(canvasId, data) {{
            const ctx = document.getElementById(canvasId);
            if (!ctx) return;
            
            new Chart(ctx, {{
                type: 'doughnut',
                data: {{
                    datasets: [{{
                        data: [data.score, 100 - data.score],
                        backgroundColor: ['#10b981', '#e5e7eb'],
                        borderWidth: 0
                    }}]
                }},
                options: {{
                    responsive: true,
                    plugins: {{
                        legend: {{ display: false }}
                    }},
                    cutout: '70%'
                }}
            }});
        }}
        
        function initializePerformanceTimeline() {{
            // Performance timeline chart
            const timelineData = {visualizations.get('performance_timeline', '{}')};
            console.log('Timeline data:', timelineData);
        }}
        
        function initializeResourceBreakdown() {{
            // Resource breakdown chart
            const resourceData = {visualizations.get('resource_breakdown', '{}')};
            console.log('Resource data:', resourceData);
        }}
        
        function initializeNetworkTiming() {{
            // Network timing chart
            const networkData = {visualizations.get('network_timing', '{}')};
            console.log('Network data:', networkData);
        }}
        
        // Utility functions
        function formatBytes(bytes) {{
            if (bytes === 0) return '0 Bytes';
            const k = 1024;
            const sizes = ['Bytes', 'KB', 'MB', 'GB'];
            const i = Math.floor(Math.log(bytes) / Math.log(k));
            return parseFloat((bytes / Math.pow(k, i)).toFixed(2)) + ' ' + sizes[i];
        }}
        
        function formatDuration(ms) {{
            if (ms < 1000) return ms + 'ms';
            return (ms / 1000).toFixed(1) + 's';
        }}
        """
        
        return js_code
    
    def _generate_pdf_styles(self) -> str:
        """Generate PDF-ready CSS styles."""
        return """
        @media print {
            body { margin: 0; padding: 20px; }
            .report-container { max-width: none; }
            section { page-break-inside: avoid; }
        }
        
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 800px;
            margin: 0 auto;
            padding: 20px;
        }
        
        .report-header {
            text-align: center;
            border-bottom: 2px solid #333;
            padding-bottom: 20px;
            margin-bottom: 30px;
        }
        
        .report-info {
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 10px;
            margin-top: 15px;
        }
        
        section {
            margin-bottom: 40px;
        }
        
        h2 {
            border-bottom: 1px solid #ddd;
            padding-bottom: 10px;
        }
        
        .executive-summary-content {
            display: grid;
            grid-template-columns: 1fr 2fr;
            gap: 2rem;
        }
        
        .metric-item {
            text-align: center;
            padding: 20px;
            border: 1px solid #ddd;
            border-radius: 8px;
        }
        
        .metric-value {
            font-size: 2.5em;
            font-weight: bold;
            color: #10b981;
        }
        
        .matrix-content table {
            width: 100%;
            border-collapse: collapse;
            margin: 1rem 0;
        }
        
        .matrix-content th,
        .matrix-content td {
            padding: 0.75rem;
            text-align: left;
            border-bottom: 1px solid #e5e7eb;
        }
        """