"""
Excel and Data Export Module

This module provides comprehensive data export capabilities:
- Excel with multiple sheets: Summary, Raw Data, Statistics, Charts
- CSV with metadata headers (URL, timestamp, platform, schema version)
- JSON with schema versioning for reproducibility
- Markdown reports for documentation
"""

import json
import statistics
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any, Union
from dataclasses import dataclass, asdict


# Schema version for reproducibility
SCHEMA_VERSION = "2.0.0"


@dataclass
class ExportMetadata:
    """Metadata for exported data files."""
    url: str
    timestamp: str
    platform: str
    schema_version: str
    scanner_version: str = "1.0.2"
    session_name: str = ""
    device_type: str = ""
    network_condition: str = ""
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return asdict(self)
    
    def to_csv_header(self) -> str:
        """Generate CSV metadata header."""
        lines = [
            f"# URL: {self.url}",
            f"# Timestamp: {self.timestamp}",
            f"# Platform: {self.platform}",
            f"# Schema Version: {self.schema_version}",
            f"# Scanner Version: {self.scanner_version}",
            f"# Session: {self.session_name}",
            f"# Device: {self.device_type}",
            f"# Network: {self.network_condition}",
            "# ---",
        ]
        return "\n".join(lines)


class ExcelExportGenerator:
    """
    Generator for Excel exports with multiple sheets.
    
    Uses openpyxl for Excel generation when available.
    Falls back to CSV generation if openpyxl is not installed.
    """
    
    def __init__(self):
        self.has_openpyxl = self._check_openpyxl()
    
    def _check_openpyxl(self) -> bool:
        """Check if openpyxl is available."""
        try:
            import openpyxl
            return True
        except ImportError:
            return False
    
    def generate_excel(self, result, url: str, output_path: Path) -> Path:
        """
        Generate Excel file with multiple sheets.
        
        Args:
            result: Scan result
            url: URL that was scanned
            output_path: Path to save Excel file
            
        Returns:
            Path to generated file
        """
        if self.has_openpyxl:
            return self._generate_excel_with_openpyxl(result, url, output_path)
        else:
            # Fall back to CSV
            return self._generate_csv_fallback(result, url, output_path)
    
    def _generate_excel_with_openpyxl(self, result, url: str, output_path: Path) -> Path:
        """Generate Excel using openpyxl."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Sheet 1: Summary
        ws_summary = wb.create_sheet("Summary")
        self._create_summary_sheet(ws_summary, result, url)
        
        # Sheet 2: Raw Data
        ws_raw = wb.create_sheet("Raw Data")
        self._create_raw_data_sheet(ws_raw, result)
        
        # Sheet 3: Statistics
        ws_stats = wb.create_sheet("Statistics")
        self._create_statistics_sheet(ws_stats, result)
        
        # Sheet 4: Core Web Vitals
        ws_cwv = wb.create_sheet("Core Web Vitals")
        self._create_cwv_sheet(ws_cwv, result)
        
        # Sheet 5: Charts (data for charts)
        ws_charts = wb.create_sheet("Chart Data")
        self._create_chart_data_sheet(ws_charts, result)
        
        # Save
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        
        return output_path
    
    def _create_summary_sheet(self, ws, result, url: str):
        """Create summary sheet."""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Title
        ws['A1'] = "Performance Scan Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Metadata
        row = 3
        metadata = [
            ("URL", url),
            ("Platform", getattr(result, 'platform', 'unknown')),
            ("Overall Score", getattr(result.performance_matrix, 'overall_score', 0)),
            ("Scan ID", getattr(result, 'scan_id', 'N/A')),
            ("Timestamp", datetime.now().isoformat()),
            ("Schema Version", SCHEMA_VERSION),
        ]
        
        for label, value in metadata:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Summary metrics
        row += 2
        ws[f'A{row}'] = "Scenario Summary"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        # Headers
        headers = ["Scenario", "Score", "Load Time (s)", "Memory (MB)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        row += 1
        
        # Data
        rows = getattr(result.performance_matrix, 'rows', [])
        for r in rows:
            ws.cell(row, 1, getattr(r.scenario, 'display_name', str(r.scenario)))
            ws.cell(row, 2, getattr(r, 'performance_score', 0))
            ws.cell(row, 3, getattr(r, 'load_time_s', 0))
            ws.cell(row, 4, getattr(r, 'memory_usage_max_mb', 0))
            row += 1
        
        # Adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 20
    
    def _create_raw_data_sheet(self, ws, result):
        """Create raw data sheet with detailed metrics."""
        from openpyxl.styles import Font, PatternFill
        
        ws['A1'] = "Raw Performance Data"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Headers
        headers = [
            "Scenario", "Device", "Network", "Score", "Load Time (s)",
            "Memory Max (MB)", "FCP (ms)", "LCP (ms)", "CLS", "TTI (ms)",
            "Total Requests", "Total Size (KB)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data
        row = 4
        scenarios = getattr(result.performance_metrics, 'scenarios', {})
        for key, scenario in scenarios.items():
            ws.cell(row, 1, scenario.scenario.display_name if hasattr(scenario, 'scenario') else key)
            ws.cell(row, 2, scenario.device_type.value if hasattr(scenario, 'device_type') else 'N/A')
            ws.cell(row, 3, scenario.network_condition.value if hasattr(scenario, 'network_condition') else 'N/A')
            ws.cell(row, 4, scenario.overall_score if hasattr(scenario, 'overall_score') else 0)
            ws.cell(row, 5, scenario.test_duration_ms / 1000 if hasattr(scenario, 'test_duration_ms') else 0)
            ws.cell(row, 6, scenario.memory_metrics.peak_heap_size_mb if hasattr(scenario, 'memory_metrics') else 0)
            ws.cell(row, 7, scenario.core_web_vitals.first_contentful_paint_ms if hasattr(scenario, 'core_web_vitals') else 0)
            ws.cell(row, 8, scenario.core_web_vitals.largest_contentful_paint_ms if hasattr(scenario, 'core_web_vitals') else 0)
            ws.cell(row, 9, scenario.core_web_vitals.cumulative_layout_shift if hasattr(scenario, 'core_web_vitals') else 0)
            ws.cell(row, 10, scenario.core_web_vitals.time_to_interactive_ms if hasattr(scenario, 'core_web_vitals') else 0)
            ws.cell(row, 11, scenario.network_metrics.total_requests if hasattr(scenario, 'network_metrics') else 0)
            ws.cell(row, 12, scenario.network_metrics.total_transfer_size_kb if hasattr(scenario, 'network_metrics') else 0)
            row += 1
    
    def _create_statistics_sheet(self, ws, result):
        """Create statistics sheet with statistical analysis."""
        from openpyxl.styles import Font, PatternFill
        
        ws['A1'] = "Statistical Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        
        headers = ["Metric", "Mean", "Median", "Std Dev", "Min", "Max", "95% CI Lower", "95% CI Upper"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Calculate statistics
        scenarios = getattr(result.performance_metrics, 'scenarios', {})
        
        if scenarios:
            scores = [s.overall_score for s in scenarios.values()]
            load_times = [s.test_duration_ms / 1000 for s in scenarios.values()]
            memories = [s.memory_metrics.peak_heap_size_mb for s in scenarios.values()]
            
            stats_data = [
                ("Performance Score", scores),
                ("Load Time (s)", load_times),
                ("Memory (MB)", memories),
            ]
            
            row = 4
            for metric_name, values in stats_data:
                if values:
                    ws.cell(row, 1, metric_name)
                    ws.cell(row, 2, statistics.mean(values))
                    ws.cell(row, 3, statistics.median(values))
                    ws.cell(row, 4, statistics.stdev(values) if len(values) > 1 else 0)
                    ws.cell(row, 5, min(values))
                    ws.cell(row, 6, max(values))
                    
                    # Calculate CI
                    if len(values) > 1:
                        mean = statistics.mean(values)
                        std_err = statistics.stdev(values) / (len(values) ** 0.5)
                        margin = 1.96 * std_err  # 95% CI
                        ws.cell(row, 7, mean - margin)
                        ws.cell(row, 8, mean + margin)
                    
                    row += 1
    
    def _create_cwv_sheet(self, ws, result):
        """Create Core Web Vitals sheet."""
        from openpyxl.styles import Font, PatternFill
        
        ws['A1'] = "Core Web Vitals"
        ws['A1'].font = Font(size=14, bold=True)
        
        headers = ["Scenario", "FCP (ms)", "LCP (ms)", "CLS", "TTI (ms)", "TBT (ms)", "Speed Index"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        row = 4
        scenarios = getattr(result.performance_metrics, 'scenarios', {})
        for key, scenario in scenarios.items():
            if hasattr(scenario, 'core_web_vitals'):
                cwv = scenario.core_web_vitals
                ws.cell(row, 1, scenario.scenario.display_name if hasattr(scenario, 'scenario') else key)
                ws.cell(row, 2, cwv.first_contentful_paint_ms)
                ws.cell(row, 3, cwv.largest_contentful_paint_ms)
                ws.cell(row, 4, cwv.cumulative_layout_shift)
                ws.cell(row, 5, cwv.time_to_interactive_ms)
                ws.cell(row, 6, cwv.total_blocking_time_ms)
                ws.cell(row, 7, cwv.speed_index_ms)
                row += 1
    
    def _create_chart_data_sheet(self, ws, result):
        """Create sheet with data formatted for charting."""
        ws['A1'] = "Chart Data"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Performance by scenario
        ws['A3'] = "Performance by Scenario"
        ws['A3'].font = Font(bold=True)
        
        ws['A4'] = "Scenario"
        ws['B4'] = "Score"
        
        row = 5
        rows = getattr(result.performance_matrix, 'rows', [])
        for r in rows:
            ws.cell(row, 1, getattr(r.scenario, 'display_name', str(r.scenario)))
            ws.cell(row, 2, getattr(r, 'performance_score', 0))
            row += 1
    
    def _generate_csv_fallback(self, result, url: str, output_path: Path) -> Path:
        """Generate CSV when openpyxl is not available."""
        csv_path = output_path.with_suffix('.csv')
        
        metadata = ExportMetadata(
            url=url,
            timestamp=datetime.now().isoformat(),
            platform=str(getattr(result, 'platform', 'unknown')),
            schema_version=SCHEMA_VERSION
        )
        
        # Generate CSV content
        lines = [metadata.to_csv_header()]
        lines.append("Scenario,Score,Load Time (s),Memory (MB),LCP (ms),CLS")
        
        rows = getattr(result.performance_matrix, 'rows', [])
        for r in rows:
            scenario = getattr(r.scenario, 'display_name', str(r.scenario))
            score = getattr(r, 'performance_score', 0)
            load_time = getattr(r, 'load_time_s', 0)
            memory = getattr(r, 'memory_usage_max_mb', 0)
            lcp = getattr(r, 'largest_contentful_paint_ms', 0)
            cls = getattr(r, 'cumulative_layout_shift', 0)
            lines.append(f'"{scenario}",{score:.2f},{load_time:.2f},{memory:.2f},{lcp:.2f},{cls:.4f}')
        
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        with open(csv_path, 'w') as f:
            f.write('\n'.join(lines))
        
        return csv_path


class DataExporter:
    """
    Main exporter class for all data export formats.
    """
    
    def __init__(self):
        self.excel_generator = ExcelExportGenerator()
    
    def export_excel(self, result, url: str, output_path: Union[str, Path]) -> Path:
        """Export to Excel format."""
        return self.excel_generator.generate_excel(result, url, Path(output_path))
    
    def export_csv(self, result, url: str, output_path: Union[str, Path]) -> Path:
        """Export to CSV format with metadata headers."""
        output_path = Path(output_path)
        
        metadata = ExportMetadata(
            url=url,
            timestamp=datetime.now().isoformat(),
            platform=str(getattr(result, 'platform', 'unknown')),
            schema_version=SCHEMA_VERSION,
            session_name=getattr(result, 'scan_id', 'unknown')
        )
        
        lines = [metadata.to_csv_header()]
        lines.append("Scenario,Score,Load Time (s),Memory (MB),LCP (ms),CLS,TTI (ms),Total Requests,Total Size (KB)")
        
        rows = getattr(result.performance_matrix, 'rows', [])
        for r in rows:
            scenario = getattr(r.scenario, 'display_name', str(r.scenario))
            score = getattr(r, 'performance_score', 0)
            load_time = getattr(r, 'load_time_s', 0)
            memory = getattr(r, 'memory_usage_max_mb', 0)
            lcp = getattr(r, 'largest_contentful_paint_ms', 0)
            cls = getattr(r, 'cumulative_layout_shift', 0)
            tti = getattr(r, 'time_to_interactive_ms', 0)
            requests = getattr(r, 'total_requests', 0)
            size = getattr(r, 'total_size_kb', 0)
            lines.append(f'"{scenario}",{score:.2f},{load_time:.2f},{memory:.2f},{lcp:.2f},{cls:.4f},{tti:.2f},{requests},{size:.2f}')
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w') as f:
            f.write('\n'.join(lines))
        
        return output_path
    
    def export_json(self, result, url: str, output_path: Union[str, Path]) -> Path:
        """Export to JSON format with schema versioning."""
        output_path = Path(output_path)
        
        data = {
            "schema_version": SCHEMA_VERSION,
            "scanner_version": "1.0.2",
            "export_timestamp": datetime.now().isoformat(),
            "url": url,
            "metadata": {
                "platform": str(getattr(result, 'platform', 'unknown')),
                "scan_id": getattr(result, 'scan_id', 'unknown'),
                "overall_score": getattr(result.performance_matrix, 'overall_score', 0)
            },
            "scenarios": []
        }
        
        scenarios = getattr(result.performance_metrics, 'scenarios', {})
        for key, scenario in scenarios.items():
            data["scenarios"].append({
                "name": scenario.scenario.display_name if hasattr(scenario, 'scenario') else key,
                "overall_score": scenario.overall_score if hasattr(scenario, 'overall_score') else 0,
                "load_time_ms": scenario.test_duration_ms if hasattr(scenario, 'test_duration_ms') else 0,
                "memory_peak_mb": scenario.memory_metrics.peak_heap_size_mb if hasattr(scenario, 'memory_metrics') else 0,
                "core_web_vitals": {
                    "fcp_ms": scenario.core_web_vitals.first_contentful_paint_ms if hasattr(scenario, 'core_web_vitals') else 0,
                    "lcp_ms": scenario.core_web_vitals.largest_contentful_paint_ms if hasattr(scenario, 'core_web_vitals') else 0,
                    "cls": scenario.core_web_vitals.cumulative_layout_shift if hasattr(scenario, 'core_web_vitals') else 0,
                    "tti_ms": scenario.core_web_vitals.time_to_interactive_ms if hasattr(scenario, 'core_web_vitals') else 0
                } if hasattr(scenario, 'core_web_vitals') else {}
            })
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w') as f:
            json.dump(data, f, indent=2)
        
        return output_path
    
    def export_markdown(self, result, url: str, output_path: Union[str, Path]) -> Path:
        """Export to Markdown format for documentation."""
        output_path = Path(output_path)
        
        lines = [
            f"# Performance Report: {url}",
            "",
            f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"**Schema Version:** {SCHEMA_VERSION}",
            f"**Platform:** {getattr(result, 'platform', 'unknown')}",
            "",
            "## Executive Summary",
            "",
            f"**Overall Score:** {getattr(result.performance_matrix, 'overall_score', 0):.1f}/100",
            "",
            "## Scenario Results",
            "",
            "| Scenario | Score | Load Time (s) | Memory (MB) |",
            "|----------|-------|---------------|-------------|",
        ]
        
        rows = getattr(result.performance_matrix, 'rows', [])
        for r in rows:
            scenario = getattr(r.scenario, 'display_name', str(r.scenario))
            score = getattr(r, 'performance_score', 0)
            load_time = getattr(r, 'load_time_s', 0)
            memory = getattr(r, 'memory_usage_max_mb', 0)
            lines.append(f"| {scenario} | {score:.1f} | {load_time:.2f} | {memory:.1f} |")
        
        lines.extend([
            "",
            "## Core Web Vitals",
            "",
            "| Metric | Value | Status |",
            "|--------|-------|--------|",
        ])
        
        if rows:
            r = rows[0]
            lcp = getattr(r, 'largest_contentful_paint_ms', 0)
            cls = getattr(r, 'cumulative_layout_shift', 0)
            
            lcp_status = "✅ Pass" if lcp <= 2500 else "⚠️ Needs Improvement"
            cls_status = "✅ Pass" if cls <= 0.1 else "⚠️ Needs Improvement"
            
            lines.append(f"| LCP | {lcp:.0f}ms | {lcp_status} |")
            lines.append(f"| CLS | {cls:.3f} | {cls_status} |")
        
        lines.extend([
            "",
            "## Recommendations",
            "",
        ])
        
        recommendations = getattr(result.performance_matrix, 'key_recommendations', [])
        for i, rec in enumerate(recommendations, 1):
            title = rec.get('title', 'Optimization')
            desc = rec.get('description', '')
            severity = rec.get('severity', 'medium')
            lines.append(f"{i}. **{title}** ({severity})\n   {desc}")
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w') as f:
            f.write('\n'.join(lines))
        
        return output_path


def export_all_formats(result, url: str, output_dir: Union[str, Path],
                       session_name: str) -> Dict[str, Path]:
    """
    Export scan result to all available formats.
    
    Args:
        result: Scan result to export
        url: URL that was scanned
        output_dir: Directory to save files
        session_name: Session name for file naming
        
    Returns:
        Dictionary mapping format names to file paths
    """
    exporter = DataExporter()
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    paths = {}
    
    try:
        paths['excel'] = exporter.export_excel(result, url, output_dir / f"{session_name}_report.xlsx")
    except Exception as e:
        paths['excel'] = None
    
    paths['csv'] = exporter.export_csv(result, url, output_dir / f"{session_name}_report.csv")
    paths['json'] = exporter.export_json(result, url, output_dir / f"{session_name}_report.json")
    paths['markdown'] = exporter.export_markdown(result, url, output_dir / f"{session_name}_report.md")
    
    return paths
